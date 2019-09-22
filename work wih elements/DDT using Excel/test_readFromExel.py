import openpyxl


def test_read_from_exel():
    fNAmeORPath = r"test.xlsx"
    workbook = openpyxl.load_workbook(fNAmeORPath)
    sheet=workbook.active
    rows = sheet.max_row
    cols = sheet.max_column

    print('Rows : {}, Collumns: {}'.format(rows,cols))

    for r in range(1,rows+1):
        for c in range(1,cols+1):
            print(sheet.cell(row=r,column=c).value,end="    ")
        print()
    
